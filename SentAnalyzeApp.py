import streamlit as st
from googleapiclient.discovery import build
import pandas as pd
import base64
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import re
import emoji
import nltk
import string
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
import pickle
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix
from sklearn.feature_extraction.text import TfidfVectorizer

# ===============================CRAWLING YOUTUBE==========================
# Fungsi untuk crawling data komentar dari video YouTube


def video_comments(api_key, video_id):
    replies = []
    youtube = build('youtube', 'v3', developerKey=api_key)
    video_response = youtube.commentThreads().list(
        part='snippet,replies', videoId=video_id).execute()

    while video_response:
        for item in video_response['items']:
            published = item['snippet']['topLevelComment']['snippet']['publishedAt']
            user = item['snippet']['topLevelComment']['snippet']['authorDisplayName']
            comment = item['snippet']['topLevelComment']['snippet']['textDisplay']
            replies.append([published, user, comment])

            replycount = item['snippet']['totalReplyCount']
            if replycount > 0:
                for reply in item['replies']['comments']:
                    published = reply['snippet']['publishedAt']
                    user = reply['snippet']['authorDisplayName']
                    repl = reply['snippet']['textDisplay']
                    replies.append([published, user, repl])

        if 'nextPageToken' in video_response:
            video_response = youtube.commentThreads().list(
                part='snippet,replies',
                pageToken=video_response['nextPageToken'],
                videoId=video_id
            ).execute()
        else:
            break

    return replies

# Fungsi untuk menampilkan halaman crawling


def show_crawling_page():
    st.header("Crawling Youtube Comments")
    api_key = st.text_input("Input API Key YouTube")
    video_ids = st.text_area(
        "Input ID Video YouTube (pisahkan dengan koma atau newline)")

    if st.button("Crawl Comments"):
        if api_key and video_ids:
            video_ids_list = [vid.strip()
                              for vid in video_ids.replace('\n', ',').split(',')]
            all_comments = []

            for vid in video_ids_list:
                comments = video_comments(api_key, vid)
                all_comments.extend(comments)

            # Menampilkan komentar dalam DataFrame
            df = pd.DataFrame(all_comments, columns=[
                              'Create-At', 'From-User', 'Text'])
            st.dataframe(df)

            # Menambahkan tombol download Excel
            st.markdown("<h3>Download Dataset</h3>", unsafe_allow_html=True)
            excel_data = [df.columns.tolist()] + df.values.tolist()
            wb = Workbook()
            ws = wb.active
            for row in excel_data:
                ws.append(row)
            excel_file = "Crawling-Result.xlsx"
            wb.save(excel_file)
            with open(excel_file, "rb") as file:
                b64 = base64.b64encode(file.read()).decode()
                href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{excel_file}">Download Excel</a>'
                st.markdown(href, unsafe_allow_html=True)
        else:
            st.text("Please input API Key and ID Video YouTube")


# ==================================PREPROCESSING==========================
# Fungsi untuk cleansing teks
def cleansing(Text):
    Text = re.sub(r'RT', '', Text)  # remove RT
    Text = Text.replace("<br>", " ")  # remove <br>
    Text = ' '.join(
        re.sub("([@#][A-Za-z0-9]+)|(\w+:\/\/\S+)", "", Text).split())
    Text = Text.lower()  # Mengubah menjadi huruf kecil
    Text = re.sub(r"[^a-zA-Z0-9]", " ", Text)
    Text = emoji.demojize(Text)  # Menghilangkan emoji
    Text = re.sub(r':[a-zA-Z_]+:', '', Text)
    Text = re.sub(r'[^\w\s]', '', Text)  # remove tanda baca
    Text = re.sub(r'\d+', '', Text)  # remove angka
    Text = Text.replace("http://", " ").replace("https://", " ")  # remove http
    Text = Text.replace('\\t', " ").replace(
        '\\n', " ").replace('\\u', " ").replace('\\', " ")
    return Text

# Fungsi untuk normalisasi teks


def slang_normalization(text):
    df_slang = pd.read_excel("normalisasi.xlsx")
    slang_dict = dict(zip(df_slang['original'], df_slang['replacement']))
    text = ' '.join(
        [slang_dict[word] if word in slang_dict else word for word in text.split()])
    return text

# Fungsi untuk tokenize teks


def tokenization(Text):
    tokens = word_tokenize(Text)
    return tokens

# Fungsi untuk tokenize teks


def remove_stopwords(tokens):
    list_stopwords = nltk.corpus.stopwords.words('indonesian')
    list_stopwords.extend(['yg', 'dg', 'dgn', 'ny', 'd', 'u', 'klo',
                           'kalo', 'amp', 'biar', 'bikin', 'bilang',
                           'gak', 'ga', 'krn', 'nya', 'nih', 'sih',
                           'si', 'tau', 'tdk', 'tuh', 'utk', 'ya',
                           'jd', 'jgn', 'sdh', 'aja', 'n', 't', 'p', 'ak',
                           'nyg', 'hehe', 'pen', 'u', 'nan', 'loh', 'rt',
                           '&amp', 'yah', 'ri', 'dci', 'di', 'iims', 'ge',
                           'eeehhhh', 'cman', 'pj', 'kyk', 'jrg',
                           'nyahnyoh', 'kya', 'hp', 'jm', 'n', 'ny',
                           'ama', 'halah', 'entut', 'drun', 'yook', 'dkk', 'a', 'lg',
                           'rd', 'do', 'aq', 'woee', 'q', 'ha', 'brow', 'de',
                           'kq', 'imho', 'hmm', 'ssh', 'aa', 'e', 'tx', 'i', 'iot',
                           'mr', 'co', 'rd', 'dr', 'imho', 'bb', 'eh', 'kl', 'koq', 'ati', 'mw',
                           'lo', 'b', 'pt', 'up', 'aaamiin', 'aama', 'aat', 'zhejiang', 'zzxxx',
                           'zack', 'zarka', 'zimbabwe', 'abang', 'abb', 'abbas', 'abisin',
                           'abrek', 'yudha', 'yuhuu', 'yupss', 'yurioshi', 'yusuf', 'ac',
                           'abrik', 'abu', 'yosua', 'your', 'youtube', 'youtubee', 'yuan', 'mudahhan'])
    txt_stopword = pd.read_csv("stopwords.txt", names=[
                               "stopwords"], header=None)
    list_stopwords.extend(txt_stopword["stopwords"][0].split(' '))
    list_stopwords = set(list_stopwords)
    tokens = [word for word in tokens if not word in list_stopwords]
    return tokens

# Fungsi untuk stemming teks


def stemming(Text):
    factory = StemmerFactory()
    stemmer = factory.create_stemmer()
    Text = [stemmer.stem(word) for word in Text]
    return Text

# Fungsi untuk menghapus tanda baca


def remove_punct(Text):
    Text = " ".join([char for char in Text if char not in string.punctuation])
    return Text

# Fungsi untuk menampilkan halaman preprocessing


def show_preprocessing_page():
    st.header("Preprocessing Data")

    # Tambahkan pilihan untuk jenis dataset
    dataset_type = st.radio("Pilih jenis dataset:", [
                            "Dengan Kolom Label", "Tanpa Kolom Label"])

    uploaded_file = st.file_uploader(
        "Upload file dataset Excel yang memiliki label sentimen", type=["xlsx"])

    if uploaded_file is not None:
        if dataset_type == "Dengan Kolom Label":
            df = pd.read_excel(uploaded_file, usecols=[
                               "From-User", "Text", "Label"])
        else:
            df = pd.read_excel(uploaded_file, usecols=["From-User", "Text"])
            df["Label"] = None  # Tambahkan kolom Label yang kosong

        st.dataframe(df)

        # Remove Netral Label
        df = df[df['Label'] != 'Netral']
        df = df[df['Label'] != 'Netral'].copy()

        # Preprocessing data dengan fungsi cleansing
        df['Cleansing'] = df['Text'].apply(cleansing)

        # Preprocessing data dengan fungsi normalization
        df['Normalization'] = df['Cleansing'].apply(slang_normalization)

        # Tambahkan proses tokenizing setelah normalisasi
        df['Tokenize'] = df['Normalization'].apply(tokenization)

        # Tambahkan proses stopwords setelah tokenizing
        df['Stopwords_Removed'] = df['Tokenize'].apply(remove_stopwords)

        # Tambahkan proses stemming setelah stopwords
        df['Stemming'] = df['Stopwords_Removed'].apply(stemming)

        # Drop duplicate
        df.drop_duplicates(subset='Stemming', keep='first', inplace=True)

        # Reset the index after dropping duplicates
        df = df.reset_index(drop=True)

        # Tambahkan proses penghapusan tanda baca setelah stemming
        df['Comment'] = df['Stemming'].apply(remove_punct)

        # Menampilkan data setelah preprocessing
        st.subheader("Cleansing Result")
        st.dataframe(df[['Cleansing', 'Label']])

        st.subheader("Normalization Result")
        st.dataframe(df[['Normalization', 'Label']])

        st.subheader("Tokenize Result")
        st.dataframe(df[['Tokenize', 'Label']])

        st.subheader("Stopwords Removed Result")
        st.dataframe(df[['Stopwords_Removed', 'Label']])

        st.subheader("Stemming Result")
        st.dataframe(df[['Stemming', 'Label']])

        st.subheader("Preprocessing Result")
        st.dataframe(df[['Comment', 'Label']])

        # Menambahkan tombol download Excel untuk Comment Result
        st.markdown("<h3>Download Preprocessing Result</h3>",
                    unsafe_allow_html=True)
        excel_data = [df[['Comment', 'Label']].columns.tolist()] + \
            df[['Comment', 'Label']].values.tolist()
        wb = Workbook()
        ws = wb.active
        for row in excel_data:
            ws.append(row)
        excel_file = "Preprocessed-Comment-Result.xlsx"
        wb.save(excel_file)
        with open(excel_file, "rb") as file:
            b64 = base64.b64encode(file.read()).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{excel_file}">Download Excel</a>'
            st.markdown(href, unsafe_allow_html=True)


# =================================CLASSIFICATION==========================
# Load model TF-IDF vectorizer, model Naive Bayes, dan model SVM
with open('tfidf_vectorizer.pkl', 'rb') as f:
    tfidf_vec_loaded = pickle.load(f)

with open('naive_bayes_tfidf.pkl', 'rb') as f:
    modelNBC_loaded = pickle.load(f)

# with open('count_vectorizer.pkl', 'rb') as f:
#     count_vec_loaded = pickle.load(f)

# with open('naive_bayes_vec.pkl', 'rb') as f:
#     modelNBC_vec_loaded = pickle.load(f)


def predict_sentiment(comment_text):
    test_tfidf_load = tfidf_vec_loaded.transform([comment_text]).toarray()
    predictNBC_load = modelNBC_loaded.predict(test_tfidf_load)
    predicted_label_NBC_load = pd.Series(
        predictNBC_load).map({0: 'Negatif', 1: 'Positif'})
    return predicted_label_NBC_load[0]

# Fungsi untuk menampilkan halaman classification


def show_classification_page():
    st.header("Classification Naive Bayes")
    st.write("Upload file dataset Test Format Excel untuk melakukan analisis sentimen")

    uploaded_file = st.file_uploader("Pilih file Excel", type=["xlsx"])

    if uploaded_file is not None:
        st.write("Ekstraksi Fitur TF-IDF:")
        test_df_load = pd.read_excel(uploaded_file)

        # Mengonversi teks pada data test ke dalam representasi TF-IDF
        comment_text = test_df_load['Comment']
        test_tfidf_load = tfidf_vec_loaded.transform(comment_text).toarray()

        df_test_tfidf_load = pd.DataFrame(
            test_tfidf_load, columns=tfidf_vec_loaded.get_feature_names_out())
        st.dataframe(df_test_tfidf_load)

        predictNBC_load = modelNBC_loaded.predict(test_tfidf_load)

        # Mengubah representasi angka kembali ke label asli
        predicted_labels_NBC_load = pd.Series(
            predictNBC_load).map({0: 'Negatif', 1: 'Positif'})

        # Membuat DataFrame hasil prediksi
        result_df_NBC_load = pd.DataFrame({
            'Comment': comment_text,
            'Label': test_df_load['Label'],
            'Classification': predicted_labels_NBC_load
        })

        st.write("Hasil Analisis Sentimen:")
        st.write(result_df_NBC_load)

        # Plot Grafik Hasil Analisis Naive Bayes
        s = pd.value_counts(result_df_NBC_load['Classification'])
        fig, ax = plt.subplots()
        ax.bar(s.index, s)
        n = len(result_df_NBC_load.index)
        for p in ax.patches:
            ax.annotate(str(round(p.get_height() / n * 100, 2)) +
                        '%', (p.get_x() * 1.005, p.get_height() * 1.005))

        plt.xlabel('Kategori')
        plt.ylabel('Jumlah')
        plt.title('Hasil Analisis Sentimen')
        st.pyplot(fig)

        # Menghitung akurasi
        label_mapping = {'Negatif': 0, 'Positif': 1}
        test_labels = test_df_load['Label'].map(label_mapping)
        accuracy = accuracy_score(test_labels, predictNBC_load)
        st.write("Akurasi:", accuracy)

        # Menampilkan confusion matrix
        matrix = confusion_matrix(test_labels, predictNBC_load)
        st.write("Confusion Matrix:")
        st.write(matrix)

        # Visualisasi confusion matrix
        labels = np.unique(test_df_load['Label'])
        sns.heatmap(matrix, annot=True, fmt='d', cmap='Blues',
                    xticklabels=labels, yticklabels=labels)
        plt.xlabel('Predicted')
        plt.ylabel('True')
        plt.title('Confusion Matrix')
        st.pyplot(fig)

        classification_report_df = pd.DataFrame(classification_report(test_labels, predictNBC_load,
                                                                      target_names=[
                                                                          'Negatif', 'Positif'],
                                                                      output_dict=True)).transpose()

        # Melakukan pembulatan pada nilai-nilai dalam tabel
        classification_report_df = classification_report_df.applymap(
            lambda x: round(x, 2) if isinstance(x, (float, int)) else x)

        st.write("Classification Report:")
        st.table(classification_report_df)

        # Menambahkan tombol download Excel untuk Comment Result
        st.markdown("<h3>Download Analysis Result</h3>",
                    unsafe_allow_html=True)
        excel_data = [result_df_NBC_load[['Comment', 'Label', 'Classification']].columns.tolist(
        )] + result_df_NBC_load[['Comment', 'Label', 'Classification']].values.tolist()
        wb = Workbook()
        ws = wb.active
        for row in excel_data:
            ws.append(row)
        excel_file = "Sentiment-Analalisis-Result.xlsx"
        wb.save(excel_file)
        with open(excel_file, "rb") as file:
            b64 = base64.b64encode(file.read()).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{excel_file}">Download Excel</a>'
            st.markdown(href, unsafe_allow_html=True)

# ==============================TAMPILAN APLIKASI==========================


def main():

    # Menambahkan sidebar untuk pilihan menu
    st.sidebar.title("Sentiment Analyze App")
    menu_options = ["Crawling", "Preprocessing", "Classification"]
    menu_choice = st.sidebar.selectbox("Pilihan Menu", menu_options)

    # Menjalankan fungsi berdasarkan pilihan menu
    if menu_choice == "Crawling":
        show_crawling_page()
    elif menu_choice == "Preprocessing":
        show_preprocessing_page()
    elif menu_choice == "Classification":
        show_classification_page()


if __name__ == "__main__":
    main()
